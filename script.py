import time
import os
import json

from fake_useragent import UserAgent

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common import NoSuchElementException
from selenium_stealth import stealth

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def get_random_chrome_user_agent():
    user_agent = UserAgent(browsers='chrome', os='windows', platforms='pc')
    return user_agent.random


def get_info():
    info = """Скрипт предназначен для парсинга отзывов на любой товар на торговой площадке Amazon.
Данные можно сохранить в форматах csv, xlsx, json."""
    return info


def get_chrome_driver():
    options = Options()
    # options.add_argument('--disable-dev-shm-usage')
    # options.add_argument('--no-sandbox')
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    # options.add_argument("--start-maximized")
    options.add_argument('--disable-gpu')
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument('--headless=new')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_experimental_option('useAutomationExtension', False)

    chromedriver = 'chromedriver.exe'
    service = ChromeService(executable_path=chromedriver)
    driver = webdriver.Chrome(options=options, service=service)

    stealth(
        driver,
        languages=["en-US", "en"],
        user_agent=get_random_chrome_user_agent(),
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
        run_on_insecure_origins=True
    )

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        'source': '''
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;
          '''
    })

    return driver


def parce_reviews():
    while True:
        link = input('Введите ссылку на товар с Amazon:\n')
        print()
        if 'https://www.amazon.co.uk/' not in link:
            print('Ссылка не правильная.')
            continue
        else:
            pages_quantity = int(input('Введите количество страниц для парсинга(не более 10ти).\n'))
            print()
            if pages_quantity > 10:
                print('Максимум 10 страниц!')
                print('Установлено значение в 10 страниц')
                pages_quantity = 10
            elif pages_quantity <= 0:
                print('Минимальное количество равно 1!')
                print('Установлено значение в 1 страницу')
                pages_quantity = 1
            print()
            print('Парсинг отзывов...')
            print()

            all_info = []
            pages_counter = 1
            browser = get_chrome_driver()
            while pages_counter <= pages_quantity:
                comms_link = link.split('/')[0] + '//' + link.split('/')[1] + link.split('/')[2] + '/' + link.split('/')[3] + '/product-reviews/' + link.split('/')[5].removesuffix('?th=1') + f'/ref=cm_cr_getr_d_paging_btm_prev_{pages_counter}?ie=UTF8&reviewerType=all_reviews&pageNumber={pages_counter}'
                browser.get(comms_link)
                page_text = browser.find_element(By.TAG_NAME, 'body').text
                if 'Enter the characters you see below' in page_text:
                    time.sleep(2)
                    browser.refresh()
                    time.sleep(2)
                try:
                    browser.find_element(By.ID, 'sp-cc-accept').click()
                except NoSuchElementException:
                    pass
                time.sleep(1)
                browser.execute_script(f"window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(1)
                soup = BeautifulSoup(browser.page_source, 'html.parser')
                pages = soup.find_all(class_='a-section review aok-relative')
                for page in pages:
                    comms = {'name': page.find(class_='a-profile-content').get_text(),
                             'date': page.find('span', {'data-hook': 'review-date'}).get_text().split('on')[1],
                             'score': page.find(class_='a-icon-alt').get_text(),
                             'country': page.find('span', {'data-hook': 'review-date'}).get_text().split('on')[
                                 0].removeprefix('Reviewed in ')}
                    try:
                        comms['title'] = page.find('a', {'data-hook': 'review-title'}).find_all('span')[-1].get_text().replace('\n', '')
                    except AttributeError:
                        comms['title'] = page.find('span', {'data-hook': 'review-title'}).get_text().replace('\n', '')
                    comms['text'] = page.find(class_='a-size-base review-text review-text-content').get_text().replace('\n', '')
                    comms['link'] = link
                    all_info.append(comms)
                pages_counter += 1

            browser.quit()
            print('Парсинг отзывов завершен.')
            print('-' * 100)

        return all_info, pages_quantity


def create_save_directory(num):
    match num:
        case '1':
            if not os.path.exists('results/csv'):
                os.makedirs('results/csv')
        case '2':
            if not os.path.exists('results/json'):
                os.makedirs('results/json')
        case '3':
            if not os.path.exists('results/excel'):
                os.makedirs('results/excel')


def result_to_csv(data):
    csv_name = input('Введите имя будущего файла.\n')
    print()

    df = pd.DataFrame.from_dict(data)
    df.to_csv(f'results/csv/{csv_name}.csv', index=False)

    print(f'Отзывы сохранены в таблице {csv_name}.csv')


def result_to_json(data):
    json_name = input('Введите имя будущего файла.\n')
    print()

    json_data = json.dumps(data)
    with open(f'results/json/{json_name}.json', 'w', encoding='utf-8') as f:
        f.write(json_data)

    print(f'Отзывы сохранены в {json_name}.json')


def result_to_excel(data, pages_quantity):
    table_name = input('Введите название будущей таблицы.\n')
    print()

    df = pd.DataFrame.from_dict(data)
    df.to_excel(f'results/excel/{table_name}.xlsx', index=False)
    workbook = load_workbook(f'results/excel/{table_name}.xlsx')
    worksheet = workbook['Sheet1']
    worksheet.merge_cells(f'G2:G{(pages_quantity * 10) + 1}')

    for column in worksheet.columns:
        column_letter = column[0].column_letter
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            worksheet.column_dimensions[column_letter].width = 20

    worksheet.column_dimensions['F'].width = 100
    workbook.save(f'results/excel/{table_name}.xlsx')

    print(f'Отзывы сохранены в таблице {table_name}.xlsx')
    print('-' * 100)


print('Скрипт для парсинга отзывов на товары с торговой площадки Amazon.')
print('-' * 100)

while True:
    print('Выберите пункт меню:')
    print('1. Спарсить отзывы на товар')
    print('2. Инфо')
    print('q. Выход')
    print('-' * 100)
    match input():
        case '1':
            rws = parce_reviews()
            while True:
                print('Выберите формат вывода данных.')
                print('1. CSV файл')
                print('2. JSON файл')
                print('3. Таблица Excel')
                print('q. Выход в главное меню')
                print('-' * 100)
                match input():
                    case '1':
                        create_save_directory('1')
                        result_to_csv(rws[0])
                    case '2':
                        create_save_directory('2')
                        result_to_json(rws[0])
                    case '3':
                        create_save_directory('3')
                        result_to_excel(rws[0], rws[1])
                    case 'q' | 'й':
                        break
                    case _:
                        print('Нет такого пункта в меню')
                        print('-' * 100)
                        continue
            print('-' * 100)
            continue
        case '2':
            print(get_info())
            print('-' * 100)
            continue
        case 'q' | 'й':
            print('Завершение работы')
            break
        case _:
            print('Нет такого пункта в меню')
            print('-' * 100)
            continue
